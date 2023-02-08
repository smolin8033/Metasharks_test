# flake8: noqa

import abc

from openpyxl.workbook import Workbook

from fields.models import Field
from groups.models import StudyGroup
from users.constants import MAX_STUDENTS_NUMBER


class ReportEntity(abc.ABC):
    @abc.abstractmethod
    def build(self):
        ...


class FieldEntity(ReportEntity):
    def build(self):
        fields = Field.objects.prefetch_related("studygroup_set").select_related("supervisors")
        fields_rows = [("Дисциплина", "Номера групп", "Ментор"), ("", "", "")]
        for field in fields:
            field_supervisor = None
            try:
                field_supervisor = field.supervisors
            except:
                ...

            fields_rows.append(
                (
                    field.name,
                    ",".join(list(map(str, field.studygroup_set.values_list("number", flat=True)))),
                    str(field_supervisor),
                )
            )
        return fields_rows


class GroupEntity(ReportEntity):
    def build(self):
        study_groups = StudyGroup.objects.prefetch_related("students")
        groups_rows = [("Номер группы", "Студенты", "Количество м/ж", "Свободных мест"), ("", "", "", "")]
        for group in study_groups:
            students = group.students.filter(role="S").order_by("last_name")
            groups_rows.append(
                (
                    group.number,
                    "\n".join(list(map(str, students.values_list("last_name", flat=True)))),
                    f"Males: {students.filter(gender='M').count()}\nFemales: {students.filter(gender='F').count()}"
                    f"\nNot given: {students.filter(gender='').count()}",
                    MAX_STUDENTS_NUMBER - len(students),
                )
            )
        return groups_rows


class ReportBuilder:
    def __init__(self, entities: list[ReportEntity]):
        self.wb = Workbook()
        self.ws = None
        self._entities = entities

    def run(self):
        for entity in self._entities:
            report_row = entity.build()
            self.make_sheet(report_row)

    def make_sheet(self, report_row: list):
        self.ws = self.wb.create_sheet()
        row = 0
        for line in report_row:
            row += 1
            for col, item in enumerate(line):
                self.ws.cell(column=col + 1, row=row, value=item)

        self.save()

    def save(self, file_name: str = "report.xlsx"):
        self.wb.save(file_name)
